# app.py
import streamlit as st
import openpyxl
import io
from database import collection

# Sayfa Yapılandırması
st.set_page_config(page_title="RPA Hakediş Asistanı", page_icon="🏗️", layout="centered")

# Başlık ve Açıklama
st.title("🏗️ Akıllı Hakediş ve Yaklaşık Maliyet Robotu")
st.markdown("""
Bu sistem, yüklediğiniz teklif cetvellerini analiz eder ve **PTT, ÇŞB, KTB** resmi fiyat veritabanlarını kullanarak orijinal Excel formatınızı bozmadan saniyeler içinde fiyatlandırır.
""")
st.divider()

# Dosya Yükleme Alanı
uploaded_file = st.file_uploader("Fiyatlandırılacak Şablonu Yükleyin (Sadece .xlsx)", type=["xlsx"])

if uploaded_file is not None:
    st.info(f"📁 Dosya hafızaya alındı: {uploaded_file.name}")
    
    # İşlemi tetikleyen buton
    if st.button("🚀 Otomatik Fiyatlandırmayı Başlat", use_container_width=True):
        
        with st.spinner('Yapay Zeka veritabanı taranıyor ve matris işleniyor... Lütfen bekleyin.'):
            
            # Excel'i RAM üzerinde açıyoruz
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active
            
            # Koordinat Tespiti (Heuristic)
            baslik_satiri, col_poz, col_fiyat, col_miktar, col_tutar = None, None, None, None, None
            for row in range(1, 25):
                for col in range(1, 15):
                    hucre_degeri = str(ws.cell(row=row, column=col).value).strip() #type: ignore
                    if "İş Kalemi No" in hucre_degeri or "Poz No" in hucre_degeri:
                        baslik_satiri, col_poz = row, col
                    elif "Birim Fiyat" in hucre_degeri:
                        col_fiyat = col
                    elif "Miktar" in hucre_degeri:
                        col_miktar = col
                    elif "Tutar" in hucre_degeri:
                        col_tutar = col
                if baslik_satiri:
                    break
            
            if not all([baslik_satiri, col_poz, col_fiyat]):
                st.error("❌ Hata: Excel şablonunda 'İş Kalemi No' veya 'Birim Fiyat' başlıkları saptanamadı.")
            else:
                basarili_eslesme = 0
                bos_birakilanlar = 0
                anomali_listesi = []
                
                # Fiyatlandırma Döngüsü
                for row in range(baslik_satiri + 1, ws.max_row + 1): #type: ignore
                    poz_hucre = ws.cell(row=row, column=col_poz).value #type: ignore
                    if poz_hucre is None: continue
                        
                    poz_no = str(poz_hucre).strip()
                    if poz_no.lower() in ['none', '', 'nan']: continue
                        
                    # Veritabanında tam metin eşleşmesi
                    sonuc = collection.get(where={"poz_no": poz_no})
                    
                    # BİRLEŞTİRİLMİŞ HÜCRE (MERGED CELL) ZIRHI
                    try:
                        if sonuc and sonuc['metadatas'] and len(sonuc['metadatas']) > 0:
                            fiyat = float(sonuc['metadatas'][0]['fiyat']) #type: ignore
                            ws.cell(row=row, column=col_fiyat).value = fiyat #type: ignore
                            
                            if col_miktar and col_tutar:
                                miktar_val = ws.cell(row=row, column=col_miktar).value #type: ignore
                                if miktar_val is not None:
                                    try:
                                        temiz_miktar = float(str(miktar_val).replace(',', '.'))
                                        ws.cell(row=row, column=col_tutar).value = temiz_miktar * fiyat #type: ignore
                                    except ValueError:
                                        pass
                            basarili_eslesme += 1
                        else:
                            ws.cell(row=row, column=col_fiyat).value = "" #type: ignore
                            if col_tutar: ws.cell(row=row, column=col_tutar).value = "" #type: ignore
                            anomali_listesi.append(poz_no)
                            bos_birakilanlar += 1
                    except AttributeError:
                        # Eğer hücre "Birleştirilmiş (Merged)" ise 'Read-Only' hatası verir.
                        # Bu durumu yoksayarak döngünün bir sonraki satıra geçmesini sağlıyoruz.
                        continue
                
                # Çıktı dosyasını RAM'de tutmak için BytesIO kullanımı
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # Başarı mesajı ve metrikler
                st.success("✅ Fiyatlandırma başarıyla tamamlandı!")
                col1, col2 = st.columns(2)
                col1.metric("Eşleşen ve İşlenen Kalem", f"{basarili_eslesme} Adet")
                col2.metric("Bulunamayan (Boş Bırakılan)", f"{bos_birakilanlar} Adet")
                
                # İndirme Butonu
                st.download_button(
                    label="📥 Fiyatlandırılmış Dosyayı İndir",
                    data=output,
                    file_name=f"FİYATLI_{uploaded_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                # Bulunamayanları Listeleyen Akıllı Panel
                if anomali_listesi:
                    with st.expander("⚠️ Veritabanında Karşılığı Olmayan Poz Numaraları"):
                        st.write("Sistemin fiyatını bulamadığı ve güvenlik gereği boş bıraktığı pozlar:")
                        for p in anomali_listesi:
                            st.caption(f"- {p}")