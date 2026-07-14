# main.py
import os
import openpyxl
from pathlib import Path
from database import collection

def teklif_cetveli_otomasyonu(girdi_excel: str):
    print("\n" + "="*75)
    print("ROBOTİK SÜREÇ OTOMASYONU (RPA): DETERMINISTIK FİYATLANDIRMA MOTORU")
    print("="*75)
    
    if not os.path.exists(girdi_excel):
        print(f"[KRİTİK HATA] '{girdi_excel}' adlı kaynak dosya dizinde bulunamadı.")
        return

    print(f"[*] '{girdi_excel}' şablonu belleğe alınıyor... (Biçimlendirme korunacak)")
    
    # Veri bütünlüğü ve hücre stillerinin (renk, font, kenarlık) korunması için load_workbook
    wb = openpyxl.load_workbook(girdi_excel)
    ws = wb.active
    
    # Dinamik Şema Haritalandırma Koordinatları
    baslik_satiri = None
    col_poz = None
    col_fiyat = None
    col_miktar = None
    col_tutar = None

    # Buluşsal (Heuristic) yöntemle sütun koordinatlarının tespiti
    for row in range(1, 25):
        for col in range(1, 15):
            hucre_degeri = str(ws.cell(row=row, column=col).value).strip() #type: ignore
            
            if "İş Kalemi No" in hucre_degeri or "Poz No" in hucre_degeri:
                baslik_satiri = row
                col_poz = col
            elif "Birim Fiyat" in hucre_degeri:
                col_fiyat = col
            elif "Miktar" in hucre_degeri:
                col_miktar = col
            elif "Tutar" in hucre_degeri:
                col_tutar = col
                
        if baslik_satiri:
            break

    if not all([baslik_satiri, col_poz, col_fiyat]):
        print("[-] Şema Uyuşmazlığı: Excel matrisinde 'İş Kalemi No' veya 'Birim Fiyat' saptanamadı.")
        return
        
    print(f"[*] Matris Algılandı: Başlık satırı indeks {baslik_satiri} üzerinde çözümlendi.")
    print("[*] Fiyatlandırma ve deterministik tarama döngüsü tetikleniyor...\n")

    basarili_eslesme = 0
    bos_birakilanlar = 0
    anomali_listesi = []

    # Başlık satırının hemen altından başlayarak son satıra kadar doğrusal tarama (O(N))
    for row in range(baslik_satiri + 1, ws.max_row + 1): #type: ignore
        poz_hucre = ws.cell(row=row, column=col_poz).value #type: ignore
        if poz_hucre is None:
            continue
            
        poz_no = str(poz_hucre).strip()
        if poz_no.lower() in ['none', '', 'nan']:
            continue
            
        # ChromaDB Vektör hafızası üzerinden tam metin (Exact Metadata Match) sorgusu
        sonuc = collection.get(where={"poz_no": poz_no})
        
        if sonuc and sonuc['metadatas'] and len(sonuc['metadatas']) > 0:
            # Veritabanında kayıtlı resmi fiyatı ve kurumu çek
            fiyat = float(sonuc['metadatas'][0]['fiyat']) #type: ignore
            kurum = sonuc['metadatas'][0].get('kurum', 'Bilinmiyor')
            
            # Değeri ilgili hücreye enjekte et
            ws.cell(row=row, column=col_fiyat).value = fiyat #type: ignore
            
            # Tutar hesaplama katmanı: Eğer Miktar mevcutsa matematiksel çarpımı işlet
            if col_miktar and col_tutar:
                miktar_val = ws.cell(row=row, column=col_miktar).value #type: ignore
                if miktar_val is not None:
                    try:
                        # Olası metinsel sayı formatlarını temizleme
                        temiz_miktar = float(str(miktar_val).replace(',', '.'))
                        ws.cell(row=row, column=col_tutar).value = temiz_miktar * fiyat #type: ignore
                    except ValueError:
                        pass
                        
            print(f"  -> [EŞLEŞTİ] Satır {row}: {poz_no} | {kurum} | {fiyat:.2f} TL")
            basarili_eslesme += 1
        else:
            # Karşılığı bulunamayan pozlar için hücreler kesinlikle boş bırakılır, manipülasyon yapılmaz
            ws.cell(row=row, column=col_fiyat).value = "" #type: ignore
            if col_tutar:
                ws.cell(row=row, column=col_tutar).value = "" #type: ignore
                
            print(f"  -> [BOŞ BIRAKILDI] Satır {row}: {poz_no} veritabanında mevcut değil.")
            anomali_listesi.append((row, poz_no))
            bos_birakilanlar += 1

    # Çıktı dosyasının üretimi
    cikti_adi = f"FİYATLANDIRILMIŞ_{girdi_excel}"
    wb.save(cikti_adi)
    
    print("\n" + "="*75)
    print(f">>> SÜREÇ TAMAMLANDI <<<")
    print(f" -> Veritabanından Çekilerek İndekslenen Satır : {basarili_eslesme}")
    print(f" -> Tam Karşılığı Olmadığı İçin Boş Bırakılan   : {bos_birakilanlar}")
    
    if anomali_listesi:
        print("\n[ANOMALİ RAPORU - MANUEL İNCELEME GEREKEN SATIRLAR]:")
        for satir, p_no in anomali_listesi:
            print(f"  * Satır {satir}: Poz No '{p_no}' sistem kütüphanesinde bulunamadı.")
            
    print(f"\n>>> Yapılandırılmış Yeni Excel Dosyası Mühürlendi: {cikti_adi}")
    print("="*75 + "\n")

if __name__ == "__main__":
    # Ana dizindeki Teklif Cetveli dosyasının adı
    target_file = "TEKLİF CETVELİ.xlsx"
    teklif_cetveli_otomasyonu(target_file)